import os
import pandas as pd
from openpyxl import load_workbook
from tkinter import messagebox

config_file = "DPWDS.cfg"

def excel_reader(f_path):
    try:
        book = load_workbook(f_path, read_only=True)
        df = pd.read_excel(f_path, sheet_name=book.sheetnames[0])
        return df
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read Excel file: {e}")
        return None

def write_excel(df, f_path):
    try:
        book = load_workbook(f_path)
        df.to_excel(f_path, index=False, sheet_name=book.sheetnames[0])
        messagebox.showinfo("Successfully", f"Data written to {f_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to write Excel file: {e}")
        return 0

def read_default_passwd():
    if os.path.exists(config_file):
        with open(config_file, 'r') as file:
            for line in file:
                if line.startswith('default_passwd='):
                    return line.split('=')[1].strip()
    return 'root'

def write_default_passwd(new_passwd):
    with open(config_file, 'w') as file:
        file.write(f'default_passwd={new_passwd}\n')

def generate_name_file(base, number):
    return f"{base}-{number:03}.n"

def get_next_number(base):
    max_number = 0
    for filename in os.listdir('.'):
        if filename.startswith(base) and filename.endswith('.n'):
            try:
                num = int(filename[len(base)+1:-2])
                if num > max_number:
                    max_number = num
            except ValueError:
                continue
    return max_number + 1

